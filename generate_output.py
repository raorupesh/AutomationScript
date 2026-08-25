"""Read a purchase-order text file, validate its line items, and write the
result into a copy of the blank Excel template.

Usage:
    python generate_output.py [input.txt] [template.xlsx] [output.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from po_parser import parse_purchase_order

MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# 1-indexed column positions in output_template_blank.xlsx, so a failed
# validation can fill in the matching cell.
COL_EXT_COST = 17  # "EXT Cost" header
COL_EXT_QTY = 20  # "EXT QTY" header

DEFAULT_INPUT = Path("purchase_order_sample.txt")
DEFAULT_TEMPLATE = Path("output/output_template_blank.xlsx")
DEFAULT_OUTPUT = Path("output/output.xlsx")


def build_row(po, item, filename: str) -> list:
    """One spreadsheet row: PO header fields followed by one line item's fields,
    in the exact column order of output_template_blank.xlsx."""
    return [
        filename,
        po.buyer,
        po.ship_terms,
        po.po_number,
        po.ref_master_po,
        po.ship_date,
        po.vendor,
        po.ship_to,
        po.bill_to,
        item.dept,
        item.sku,
        item.upc,
        item.vendor_part,
        item.description,
        item.retail,
        item.cost,
        item.ext_cost,
        item.ctns,
        item.cspk,
        item.ext_qty,
        item.cube,
        item.kilograms,
        round(item.computed_ext_cost, 2),
    ]


def main() -> None:
    args = sys.argv[1:]
    input_path = Path(args[0]) if len(args) > 0 else DEFAULT_INPUT
    template_path = Path(args[1]) if len(args) > 1 else DEFAULT_TEMPLATE
    output_path = Path(args[2]) if len(args) > 2 else DEFAULT_OUTPUT

    text = input_path.read_text(encoding="utf-8")
    po = parse_purchase_order(text)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    mismatches = []
    for item in po.items:
        ws.append(build_row(po, item, input_path.name))
        row_num = ws.max_row
        if not item.ext_qty_ok:
            ws.cell(row=row_num, column=COL_EXT_QTY).fill = MISMATCH_FILL
            mismatches.append((item.sku, "EXT QTY", item.ext_qty, item.computed_ext_qty))
        if not item.ext_cost_ok:
            ws.cell(row=row_num, column=COL_EXT_COST).fill = MISMATCH_FILL
            mismatches.append((item.sku, "EXT COST", item.ext_cost, round(item.computed_ext_cost, 2)))

    # output_template_blank.xlsx's own columns are the required schema, so
    # mismatches are flagged in-place (red fill) rather than by adding a
    # column. A second sheet lists them explicitly for anything downstream
    # that just reads cell values.
    val_ws = wb.create_sheet("Validation")
    val_ws.append(["SKU", "Field", "Printed on PO", "Recomputed"])
    for sku, field_name, printed, computed in mismatches:
        val_ws.append([sku, field_name, printed, computed])
    if not mismatches:
        val_ws.append(["No discrepancies - every EXT QTY / EXT COST matched cartons x case pack / cost x ext qty."])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Parsed {len(po.items)} line item(s) from {input_path.name}")
    print(f"Wrote {len(po.items)} row(s) to {output_path}")
    if mismatches:
        print(f"WARNING: {len(mismatches)} validation mismatch(es) - see the 'Validation' sheet in {output_path}")
    else:
        print("Validation OK: recomputed EXT QTY / EXT COST matched the printed values on every row.")


if __name__ == "__main__":
    main()
